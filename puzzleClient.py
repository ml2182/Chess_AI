import requests
import json
import openpyxl
class data_from_excel_to_database:
    def __init__(self,wb_name):
        self.__wb_name = wb_name
        self.__wb = openpyxl.load_workbook(self.__wb_name)
        self.__ws = self.__wb.active
    def upload(self,puzzle_excel):
        rating = 0
        for i in range(2,self.__ws.max_row+1):
            for j in range(1,4):
                value = [self.__ws.cell(row=i,column=j).value]
                if j == 1:
                    FEN = str(value[0])
                if j == 2:
                    moves = str(value[0])
                if j == 3:
                    rating = int(value[0])
            print(FEN,moves,rating)
            json.loads(requests.post("https://17mlee.eu.pythonanywhere.com", json={"FEN":FEN,"moves":moves,"rating":rating}).text)
Puzzle= data_from_excel_to_database(r"C:\Users\17mlee\Colyton Grammar School\Computer Science 13A-Cs 2023 CGS - 13_0 NEA\Chess_slight_improvement - Copy - Copy - Copy - Copy - Copy - Copy\Puzzledbtest.xlsx")
Puzzle.upload(r"C:\Users\17mlee\Colyton Grammar School\Computer Science 13A-Cs 2023 CGS - 13_0 NEA\Chess_slight_improvement - Copy - Copy - Copy - Copy - Copy - Copy\Puzzledbtest.xlsx")
# Puzzle.upload(r"C:\Users\17MLee\OneDrive - Colyton Grammar School\Documents\puzzledbtrial.xlsx")

